# -*- coding: utf-8 -*-
import requests
import re
import time
import openpyxl
import matplotlib.pyplot as plt
from matplotlib.pyplot import MultipleLocator
import smtplib
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from email.mime.image import MIMEImage

your_name = '统一认证账号'
your_pwd = '统一认证密码'
dir_name = 'C:\\Users\\xxx\\xxx\\xxx\\' #存放数据文件的地址 注意要使用绝对路径
file_name=dir_name+time.strftime("%m-%d", time.localtime())
dataBase_name = 'dateBase.xlsx'

msg_from = 'xxx@qq.com'  # 发送方邮箱
passwd = 'xxxxx'  # 填入发送方邮箱的授权码
msg_to = 'xxxx@qq.com'  # 收件人邮箱


def buaaLogin(user_name, password):
    print("统一认证登录")

    postUrl = "https://app.buaa.edu.cn/uc/wap/login/check"
    postData = {
        "username": user_name,
        "password": password,
    }
    responseRes = requests.post(postUrl, data=postData)
    # 无论是否登录成功，状态码一般都是 statusCode = 200
    # print(f"statusCode = {responseRes.status_code}")
    print(f"text = {responseRes.text}")
    # print(responseRes.headers['set-cookie'])
    return responseRes


def sendEmail(res):
    s = requests.session()
    date = time.strftime("%Y-%m-%d", time.localtime())
    headers = {
        'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,image/webp,image/apng,*/*;q=0.8,application/signed-exchange;v=b3;q=0.9',
        'Referer': 'https://app.buaa.edu.cn/site/epideAll/epideType?type=weishangbao&date='+date+'&group_id=1&group_type=1&flag=2',
        'X-Requested-With': 'XMLHttpRequest',
        'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/84.0.4147.105 Safari/537.36',
        'Content-Type': 'application/x-www-form-urlencoded',
        'Cookie': res.headers['set-cookie']
    }
    r = s.get('https://app.buaa.edu.cn/xisuncov/wap/buaa/export-download?group_id=1&group_type=1&type=weishangbao&date='+date+'&flag=2', headers=headers)
    print(r)
    return r

def dataBase():
    wb1 = openpyxl.load_workbook(file_name+'.xlsx')
    ws1 = wb1['Sheet']
    wb2 = openpyxl.load_workbook(dir_name+dataBase_name)
    ws2 = wb2['Sheet']
    today =[]
    dB = []
    for cell in ws1.iter_cols(1, 2, values_only=True):
        today = cell[1:]
    for cell in ws2.iter_cols(1, 2, values_only=True):
        dB = cell[1:]

    for i in today:
        if i in dB:
            index2 = dB.index(i)+2
            rank = ws2.cell(index2,3).value
            ws2.cell(index2,3).value = rank+1
        else:
            index1 = today.index(i)+2
            ws2.append([ws1.cell(index1,1).value, i, 1])
    wb2.save(dataBase_name)
    print(file_name+"信息已更新")
    return len(today)

def draw():
    wb2 = openpyxl.load_workbook(dir_name+dataBase_name)
    ws2 = wb2['Sheet']
    name = ()
    cout = ()
    for cell in ws2.iter_cols(0, 1, values_only=True):
        name = cell[1:]
    for cell in ws2.iter_cols(2, 3, values_only=True):
        cout = cell[1:]

    plt.rcParams['font.sans-serif'] = ['SimHei']  # 显示中文标签
    plt.rcParams['axes.unicode_minus'] = False  # 这两行需要手动设置
    plt.tick_params(labelsize=8)
    plt.figure(figsize=(len(name), max(cout)+1))
    plt.plot(name, cout, 'g*-')
    for i in range(0,len(name)):
        plt.text(name[i], cout[i], cout[i], fontdict={'fontsize': 16, 'color': 'red'})
    plt.title('截止到'+file_name[45:]+'未打卡人数统计')

    plt.savefig(file_name+'.jpg')
    print("已保存到"+file_name+".jpg")
    # plt.show()

def sendPic():

    subject = "截止"+file_name[45:]+"未填报统计"  # 主题
    msg = MIMEMultipart('related')
    content = MIMEText('<html><body><img src="cid:imageid" alt="imageid"></body></html>', 'html', 'utf-8')  # 正文
    # msg = MIMEText(content)
    msg.attach(content)
    msg['Subject'] = subject
    msg['From'] = msg_from
    msg['To'] = msg_to

    file = open(file_name+".jpg", "rb")
    img_data = file.read()
    file.close()

    img = MIMEImage(img_data)
    img.add_header('Content-ID', 'imageid')
    msg.attach(img)

    try:
        s = smtplib.SMTP_SSL("smtp.qq.com", 465)  # 邮件服务器及端口号
        s.login(msg_from, passwd)
        s.sendmail(msg_from, msg_to, msg.as_string())
    except:
        print("发送到邮箱失败")
    finally:
        print("成功发送到邮箱")
        s.quit()


if __name__ == "__main__":

    #从返回结果来看，有登录成功
    res = sendEmail(buaaLogin(your_name, your_pwd))
    with open(file_name+".xlsx","wb") as file:
        file.write(res.content)
    print("今日有{}人未打卡".format(dataBase()))
    draw()
    sendPic()
